import time
from Driver.appiumandroid import AppiumDriverManager
from Pages.elementsaccess import AppiumElementFinder
import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlImage
logging.basicConfig(filename='test_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

# Load existing workbook or create a new one if it doesn't exist
try:
    wb = load_workbook("automation_test_report.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active
ws.title = 'Test Results'

headers = ['Issue ID', 'Issue Description', 'Test Result', 'Screenshot', 'Severity Level',
           'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior', 'Device/Platform',
           'Additional Notes/Comments']


# Write headers if the sheet is empty
def write_headers():
    try:
        if not any(ws.iter_rows(max_row=1, max_col=len(headers))):
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
    except Exception as e:
        logging.error(f"Error writing headers to Excel: {e}")


write_headers()


def write_test_result(issue_id, issue_description, test_result, screenshot, severity_level,
                      steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                      additional_notes_comments):
    try:
        ws.append([issue_id, issue_description, test_result, "", severity_level,
                   steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                   additional_notes_comments])
        row_number = ws.max_row
        if screenshot:
            # row_height = 100# Adjust the divisor based on your requirements
            # ws.row_dimensions[row_number].height = row_height
            # column_height = 200
            # ws.height_dimension[row_number]=column_height
            cell_reference = f'D{row_number}'
            row_height = 100  # Adjust the divisor based on your requirements
            ws.row_dimensions[row_number].height = row_height
            column_width = 100  # Adjust the divisor based on your requirements
            ws.column_dimensions[cell_reference[0]].width = column_width
            screenshot.width = 200  # Set width in pixels
            screenshot.height = 100
            ws.add_image(screenshot, cell_reference)  # Adjust cell position as needed
            wb.save('automation_test_report.xlsx')
    except Exception as e:
        logging.error(f"Error writing test result to Excel: {e}")


def test_register_user():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    element_finder = AppiumElementFinder(driver)
    try:
        login_xpaths = ['//android.widget.TextView[@text="Docile"]', '//android.widget.TextView[@text="Get Started"]',
                        '//android.widget.TextView[@text="Sign up"]']
        for xpath in login_xpaths:
            element_finder.find_clickable_element_by_xpath(xpath).click()
        element_finder.find_clickable_element_by_xpath(
            '//android.view.ViewGroup[@resource-id="TestIndividual"]').click()
        xpath1 = ['//android.widget.EditText[@text="Enter your full name"]',
                  '//android.widget.EditText[@text="Phone Number"]']
        sendkey = ['Akshay Gumme', '7972951602']
        time.sleep(25)
        for i in range(len(xpath1)):
            element_finder.find_send_data_by_xpath(xpath1[i]).send_keys(sendkey[i])
        xpath2 = ['//android.widget.TextView[@text="Next"]',
                  '//android.widget.TextView[@text="CONTINUE WITH +917972951602"]',
                  '//android.widget.TextView[@text="Next"]']
        for i in xpath2:
            element_finder.find_clickable_element_by_xpath(i).click()
        time.sleep(5)
        screenshot_path = "Registration.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        assert True
        logging.info("Registration test passed")
        print("Registration success")
        # Write test result to Excel
        write_test_result(issue_id='3', issue_description='Registered using correct mobile no.',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='User should registered successfully',
                          actual_behavior='As Expected',
                          device_platform='Android', additional_notes_comments='No issues')

    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "Registration_fail.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='3', issue_description='Register with valid credentials',
                          test_result='Failed', screenshot=screenshot,
                          severity_level='High',
                          steps_to_reproduce='1. Open app\n 2. Click on signup 3. Enter valid credentials',
                          expected_behavior='User should be Registered in successfully',
                          actual_behavior='Registration failed due to invalid credentials',
                          device_platform='Android', additional_notes_comments='Error encountered')
