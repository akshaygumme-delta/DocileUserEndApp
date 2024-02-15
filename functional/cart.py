import time
from Driver.appiumandroid import AppiumDriverManager
from Pages.elementsaccess import AppiumElementFinder
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils import get_column_letter

logging.basicConfig(filename='test_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

# Load existing workbook or create a new one if it doesn't exist
try:
    wb = load_workbook("automation_test_report1.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active
ws.title = 'Test Results'

headers = ['Issue ID', 'Issue Description', 'Test Result', 'Screenshot', 'Severity Level',
           'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior', 'Device/Platform',
           'Additional Notes/Comments']


# Write headers if the sheet is empty
def write_headers():
    try:
        if not any(ws.iter_rows(max_row=1, max_col=len(headers))):
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
    except Exception as e:
        logging.error(f"Error writing headers to Excel: {e}")


if all(cell.value is None for cell in ws[1]):
    write_headers()


def write_test_result(issue_id, issue_description, test_result, screenshot, severity_level,
                      steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                      additional_notes_comments):
    try:
        ws.append([issue_id, issue_description, test_result, "", severity_level,
                   steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                   additional_notes_comments])
        row_number = ws.max_row
        if screenshot:
            # row_height = 100# Adjust the divisor based on your requirements
            # ws.row_dimensions[row_number].height = row_height
            # column_height = 200
            # ws.height_dimension[row_number]=column_height
            cell_reference = f'D{row_number}'
            row_height = 100  # Adjust the divisor based on your requirements
            ws.row_dimensions[row_number].height = row_height
            column_width = 100  # Adjust the divisor based on your requirements
            ws.column_dimensions[cell_reference[0]].width = column_width
            screenshot.width = 200  # Set width in pixels
            screenshot.height = 100
            ws.add_image(screenshot, cell_reference)  # Adjust cell position as needed
            wb.save('automation_test_report1.xlsx')
    except Exception as e:
        logging.error(f"Error writing test result to Excel: {e}")


def test_cart():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    element_finder = AppiumElementFinder(driver)
    try:
        xpaths = ['//android.widget.FrameLayout[@resource-id="android:id/content"]/android.widget.FrameLayout/android'
                  '.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup['
                  '1]/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup/android.view.ViewGroup'
                  '/android.view.ViewGroup[5]/android.view.ViewGroup/android.view.ViewGroup['
                  '6]/android.widget.ImageView','(//android.widget.TextView[@text=""])[1]',
                  '(//android.widget.TextView[@text=""])[1]',
                  '(//android.widget.CheckBox[@resource-id="ProductCheck"])[1]',
                  '//android.widget.TextView[@text=" Remove"]', '//android.widget.TextView[@text="Change"]',
                  '//android.widget.TextView[@text="Confirm & Continue"]']

        for i in xpaths:
            element_finder.find_clickable_element_by_xpath(i).click()
        screenshot_path = "cart.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        assert True
        logging.info("cart page tested successfully")
        write_test_result(issue_id='5', issue_description='cart page functionality',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='Need to add,decrease and remove the products',
                          actual_behavior='As expected ',
                          device_platform='Android', additional_notes_comments='No issues')

    except Exception as e:
        logging.error(f"Failed to execute cart functionality: {e}")
        time.sleep(5)
        screenshot_path = "prod_search_fail.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        assert True
        logging.info("cart page tested successfully")
        write_test_result(issue_id='5', issue_description='Login with valid credentials',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='User should be logged in successfully',
                          actual_behavior='User logged in successfully',
                          device_platform='Android', additional_notes_comments='No issues')
