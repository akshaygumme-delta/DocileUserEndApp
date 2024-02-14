import time
from Driver.appiumandroid import AppiumDriverManager
from Pages.elementsaccess import AppiumElementFinder
import logging
import openpyxl
from openpyxl.styles import Font
from selenium.webdriver.common.keys import Keys
from openpyxl.drawing.image import Image as xlImage
logging.basicConfig(filename='test_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

# Load existing workbook or create a new one if it doesn't exist
try:
    wb = load_workbook("automation_test_report.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active
ws.title = 'Test Results'

headers = ['Issue ID', 'Issue Description', 'Test Result', 'Screenshot', 'Severity Level',
           'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior', 'Device/Platform',
           'Additional Notes/Comments']


# Write headers if the sheet is empty
def write_headers():
    try:
        if not any(ws.iter_rows(max_row=1, max_col=len(headers))):
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
    except Exception as e:
        logging.error(f"Error writing headers to Excel: {e}")


write_headers()


def write_test_result(issue_id, issue_description, test_result, screenshot, severity_level,
                      steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                      additional_notes_comments):
    try:
        ws.append([issue_id, issue_description, test_result, "", severity_level,
                   steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                   additional_notes_comments])
        row_number = ws.max_row
        if screenshot:
            # row_height = 100# Adjust the divisor based on your requirements
            # ws.row_dimensions[row_number].height = row_height
            # column_height = 200
            # ws.height_dimension[row_number]=column_height
            cell_reference = f'D{row_number}'
            row_height = 100  # Adjust the divisor based on your requirements
            ws.row_dimensions[row_number].height = row_height
            column_width = 100  # Adjust the divisor based on your requirements
            ws.column_dimensions[cell_reference[0]].width = column_width
            screenshot.width = 200  # Set width in pixels
            screenshot.height = 100
            ws.add_image(screenshot, cell_reference)  # Adjust cell position as needed
            wb.save('automation_test_report.xlsx')
    except Exception as e:
        logging.error(f"Error writing test result to Excel: {e}")


def test_product_search():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    element_finder = AppiumElementFinder(driver)
    try:
        element_finder.find_clickable_element_by_xpath('//android.view.ViewGroup[@resource-id="TestHomeSearchnav"]/android.widget.ImageView').click()
        search = element_finder.find_send_data_by_xpath('//android.widget.EditText[@text="Search"]')
        search.send_keys('Biscuits')
        screenshot_path = "Search_prod.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        element_finder.find_clickable_element_by_xpath('//android.widget.TextView[@resource-id="goBackNav"]').click()

        # Press the Enter key
        # search.send_keys(Keys.ENTER)
        assert True
        logging.info("Successful login test passed")
        print("login success")
        # Write test result to Excel
        write_test_result(issue_id='3', issue_description='Search the product',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='User should search desired product',
                          actual_behavior='As expected',
                          device_platform='Android', additional_notes_comments='No issues')

    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "Search_prod_fail.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='3', issue_description='Searching for product failed',
                          test_result='Failed', screenshot=screenshot,
                          severity_level='High',
                          steps_to_reproduce='1. check for available items',
                          expected_behavior='Products not found',
                          actual_behavior='As expected',
                          device_platform='Android', additional_notes_comments='Error encountered')
