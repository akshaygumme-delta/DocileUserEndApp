# import time
# from Driver.appiumandroid import AppiumDriverManager
# from Pages.elementsaccess import AppiumElementFinder
# import logging
# import openpyxl
# from openpyxl.styles import Font
# from openpyxl.drawing.image import Image as xlImage
# logging.basicConfig(filename='test_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
#
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet.title = 'Test Results'
#
# headers = ['Issue ID', 'Issue Description', 'Test Result', 'Screenshot', 'Severity Level',
#            'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior', 'Device/Platform',
#            'Additional Notes/Comments']
# sheet.append(headers)
#
# for col in range(1, len(headers) + 1):
#     sheet.cell(row=1, column=col).font = Font(bold=True)
#
#
# def write_test_result(issue_id, issue_description, test_result, screenshot, severity_level,
#                       steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
#                       additional_notes_comments):
#     try:
#         sheet.append([issue_id, issue_description, test_result, screenshot, severity_level,
#                       steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
#                       additional_notes_comments])
#         workbook.save('testcase_results.xlsx')
#     except Exception as e:
#         logging.error(f"Error writing test result to Excel: {e}")
#
#
# def test_successful_login():
#     driver_manager = AppiumDriverManager()
#     driver = driver_manager.appium_driver()
#     element_finder = AppiumElementFinder(driver)
#     try:
#         login_xpaths = ['//android.widget.TextView[@text="Docile"]', '//android.widget.TextView[@text="Get Started"]',
#                         '//android.widget.TextView[@text="Sign in"]']
#         for xpath in login_xpaths:
#             element_finder.find_clickable_element_by_xpath(xpath).click()
#
#         element_finder.find_send_data_by_xpath('//android.widget.EditText[@text="Phone Number"]').send_keys(
#             "7972951602")
#
#         element_finder.find_clickable_element_by_xpath('//android.widget.TextView[@text="Continue"]').click()
#         time.sleep(25)
#         element_finder.find_clickable_element_by_xpath('//android.widget.TextView[@text="Next"]').click()
#         screenshot= f"login_success.png"
#         driver.save_screenshot(screenshot)
#
#         welcome_message = element_finder.find_visible_element_by_xpath(
#             '//android.widget.TextView[@text="Docile"]').text
#         assert "Docile" in welcome_message, "Login was not successful"
#         driver.save_screenshot('/output/login_success.png')
#         logging.info("Successful login test passed")
#         print("login success")
#         # Write test result to Excel
#         write_test_result(issue_id='1', issue_description='Login with valid credentials',
#                           test_result='Passed', screenshot="",
#                           severity_level='None', steps_to_reproduce='None',
#                           expected_behavior='User should be logged in successfully',
#                           actual_behavior='User logged in successfully',
#                           device_platform='Android', additional_notes_comments='No issues')
#         img = xlImage(screenshot)
#         img.width = 3 * 96 * 0.393701
#         img.height = 6 * 96 * 0.393701
#         sheet.add_image(img, 'D2')
#         workbook.save("test_results.xlsx")
#     except Exception as e:
#         logging.error(f"Failed to execute successful login test: {e}")
#         login_fail=driver.save_screenshot('login_fail.png')
#         # Write test result to Excel
#         write_test_result(issue_id='1', issue_description='Login with valid credentials',
#                           test_result='Failed', screenshot=login_fail,
#                           severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
#                           expected_behavior='User should be logged in successfully',
#                           actual_behavior='Login failed',
#                           device_platform='Android', additional_notes_comments='Error encountered')
#

# def test_product_search():
#     driver_manager = AppiumDriverManager()
#     driver = driver_manager.appium_driver()
#     element_finder = AppiumElementFinder(driver)
#     try:
#         element_finder.find_clickable_element_by_xpath(
#             '//android.view.ViewGroup[@resource-id="TestHomeSearchnav"]/android.widget.ImageView').click()
#         search = element_finder.find_send_data_by_xpath('//android.widget.EditText[@text="Search"]')
#         search.send_keys('Biscuits')
#
#         # Press the Enter key
#         # search.send_keys(Keys.ENTER)
#         assert True
#         product_search = driver.save_screenshot('search_prod_success.png')
#         logging.info("Successful login test passed")
#         print("product searched")
#         # Write test result to Excel
#         write_test_result(issue_id='2', issue_description='Search the product',
#                           test_result='Passed', screenshot= product_search,
#                           severity_level='None', steps_to_reproduce='None',
#                           expected_behavior='User should search desired product',
#                           actual_behavior='As expected',
#                           device_platform='Android', additional_notes_comments='No issues')
#
#     except Exception as e:
#         logging.error(f"Failed to execute successful login test: {e}")
#         driver.save_screenshot('/output/prod_search_fail.png')
#         # Write test result to Excel
#         write_test_result(issue_id='2', issue_description='Searching for product failed',
#                           test_result='Failed', screenshot='/output/prod_search_fail.png',
#                           severity_level='High',
#                           steps_to_reproduce='1. check for available items',
#                           expected_behavior='Products not found',
#                           actual_behavior='As expected',
#                           device_platform='Android', additional_notes_comments='Error encountered')

#
# def test_invalid_login():
#     try:
#         driver_manager = AppiumDriverManager()
#         driver = driver_manager.appium_driver()
#         element_finder = AppiumElementFinder(driver)
#
#         login_xpaths = ['//android.widget.TextView[@text="Docile"]', '//android.widget.TextView[@text="Get Started"]',
#                         '//android.widget.TextView[@text="Sign in"]']
#         for xpath in login_xpaths:
#             element_finder.find_clickable_element_by_xpath(xpath).click()
#
#         element_finder.find_send_data_by_xpath('//android.widget.EditText[@text="Phone Number"]').send_keys("87878778**")
#         element_finder.find_clickable_element_by_xpath('//android.widget.TextView[@text="Continue"]').click()
#
#         error_message = element_finder.find_visible_element_by_xpath(
#             '//android.widget.TextView[@text="Account not found, or not activated"]').text
#         assert "not found" in error_message, "Login was expected to fail with invalid credentials"
#         logging.info("Invalid login test passed")
#
#         # Write test result to Excel
#         write_test_result(issue_id='2', issue_description='Login with invalid credentials',
#                           test_result='Passed', screenshot='screenshot2.png',
#                           severity_level='Medium', steps_to_reproduce='1. Open app\n2. Enter invalid credentials',
#                           expected_behavior='Login should fail',
#                           actual_behavior='Login failed as expected',
#                           device_platform='Android', additional_notes_comments='No issues')
#
#     except Exception as e:
#         # Log test failure
#         logging.error(f"Failed to execute invalid login test: {e}")
#
#         # Write test result to Excel
#         write_test_result(issue_id='2', issue_description='Login with invalid credentials',
#                           test_result='Failed', screenshot='screenshot2.png',
#                           severity_level='High', steps_to_reproduce='1. Open app\n2. Enter invalid credentials',
#                           expected_behavior='Login should fail',
#                           actual_behavior='Login did not fail as expected',
#                           device_platform='Android', additional_notes_comments='Error encountered')
#


import time
from Driver.appiumandroid import AppiumDriverManager
from Pages.elementsaccess import AppiumElementFinder
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils import get_column_letter

logging.basicConfig(filename='test_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

# Load existing workbook or create a new one if it doesn't exist
try:
    wb = load_workbook("automation_test_report1.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active
ws.title = 'Test Results'

headers = ['Issue ID', 'Issue Description', 'Test Result', 'Screenshot', 'Severity Level',
           'Steps to Reproduce', 'Expected Behavior', 'Actual Behavior', 'Device/Platform',
           'Additional Notes/Comments']


# Write headers if the sheet is empty
def write_headers():
    try:
        if all(cell.value is None for cell in ws[1]):
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
    except Exception as e:
        logging.error(f"Error writing headers to Excel: {e}")


write_headers()


def write_test_result(issue_id, issue_description, test_result, screenshot, severity_level,
                      steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                      additional_notes_comments):
    try:
        ws.append([issue_id, issue_description, test_result, "", severity_level,
                   steps_to_reproduce, expected_behavior, actual_behavior, device_platform,
                   additional_notes_comments])
        row_number = ws.max_row
        if screenshot:
            # row_height = 100# Adjust the divisor based on your requirements
            # ws.row_dimensions[row_number].height = row_height
            # column_height = 200
            # ws.height_dimension[row_number]=column_height
            cell_reference = f'D{row_number}'
            row_height = 100  # Adjust the divisor based on your requirements
            ws.row_dimensions[row_number].height = row_height
            column_width = 100  # Adjust the divisor based on your requirements
            ws.column_dimensions[cell_reference[0]].width = column_width
            screenshot.width = 200  # Set width in pixels
            screenshot.height = 100
            ws.add_image(screenshot, cell_reference)  # Adjust cell position as needed
            wb.save('automation_test_report1.xlsx')
    except Exception as e:
        logging.error(f"Error writing test result to Excel: {e}")


def test_successful_login():
    driver_manager = AppiumDriverManager()
    driver = driver_manager.appium_driver()
    element_finder = AppiumElementFinder(driver)
    try:
        login_xpaths = ['//android.widget.TextView[@text="Docile"]', '//android.widget.TextView[@text="Get Started"]',
                        '//android.widget.TextView[@text="Sign in"]']
        for xpath in login_xpaths:
            element_finder.find_clickable_element_by_xpath(xpath).click()

        element_finder.find_send_data_by_xpath('//android.widget.EditText[@text="Phone Number"]').send_keys(
            "7972951602")

        element_finder.find_clickable_element_by_xpath('//android.widget.TextView[@text="Continue"]').click()
        time.sleep(25)
        element_finder.find_clickable_element_by_xpath('//android.widget.TextView[@text="Next"]').click()
        screenshot_path = "login.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        welcome_message = element_finder.find_visible_element_by_xpath(
            '//android.widget.TextView[@text="Docile"]').text
        assert "Docile" in welcome_message, "Login was not successful"
        logging.info("Successful login test passed")
        print("login success")
        # Write test result to Excel
        write_test_result(issue_id='1', issue_description='Login with valid credentials',
                          test_result='Passed', screenshot=screenshot,
                          severity_level='None', steps_to_reproduce='None',
                          expected_behavior='User should be logged in successfully',
                          actual_behavior='User logged in successfully',
                          device_platform='Android', additional_notes_comments='No issues')

    except Exception as e:
        logging.error(f"Failed to execute successful login test: {e}")
        screenshot_path = "login_fail.png"
        driver.save_screenshot(screenshot_path)
        screenshot = xlImage(screenshot_path)
        # Write test result to Excel
        write_test_result(issue_id='1', issue_description='Login with valid credentials',
                          test_result='Failed', screenshot=screenshot,
                          severity_level='High', steps_to_reproduce='1. Open app\n2. Enter valid credentials',
                          expected_behavior='User should be logged in successfully',


                          actual_behavior='Login failed',
                          device_platform='Android', additional_notes_comments='Error encountered')


# def test_product_search():
#     driver_manager = AppiumDriverManager()
#     driver = driver_manager.appium_driver()
#     element_finder = AppiumElementFinder(driver)
#     try:
#         element_finder.find_clickable_element_by_xpath('//android.view.ViewGroup[@resource-id="TestHomeSearchnav"]/android.widget.ImageView').click()
#         search = element_finder.find_send_data_by_xpath('//android.widget.EditText[@text="Search"]')
#         search.send_keys('Biscuits')
#         time.sleep(5)
#         screenshot_path = "prod_search.png"
#         driver.save_screenshot(screenshot_path)
#         screenshot = xlImage(screenshot_path)
#
#         # Press the Enter key
#         # search.send_keys(Keys.ENTER)
#         assert True
#         logging.info("Successful login test passed")
#         print("login success")
#         # Write test result to Excel
#         write_test_result(issue_id='3', issue_description='Search the product',
#                           test_result='Passed', screenshot=screenshot,
#                           severity_level='None', steps_to_reproduce='None',
#                           expected_behavior='User should search desired product',
#                           actual_behavior='As expected',
#                           device_platform='Android', additional_notes_comments='No issues')
#
#     except Exception as e:
#         logging.error(f"Failed to execute successful login test: {e}")
#         time.sleep(5)
#         screenshot_path = "prod_search_fail.png"
#         driver.save_screenshot(screenshot_path)
#         screenshot = xlImage(screenshot_path)
#         # Write test result to Excel
#         write_test_result(issue_id='3', issue_description='Searching for product',
#                           test_result='Failed', screenshot=screenshot,
#                           severity_level='High',
#                           steps_to_reproduce='1. check for available items',
#                           expected_behavior='Products not found',
#                           actual_behavior='As expected',
#                           device_platform='Android', additional_notes_comments='Error encountered')
